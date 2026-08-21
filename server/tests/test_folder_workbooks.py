from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from server.app.folder_workbooks import (
    FolderExecuteRequest,
    FolderSelection,
    FolderSnapshotRequest,
    create_folder_snapshot,
    execute_folder_plan,
    refresh_folder,
    scan_folder,
)
from server.app.models import AnalysisPlan


def _stamp_plan(plan: AnalysisPlan, snapshot) -> AnalysisPlan:
    return plan.model_copy(
        update={
            "sourceFingerprint": snapshot.sourceFingerprint,
            "sourceFingerprintSheets": snapshot.sourceFingerprintSheets,
        }
    )


def _create_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "得分"
    sheet.append(["人员", "得分"])
    sheet.append(["嘟嘟嘟", 33])
    sheet.append(["阿里", 44])
    workbook.create_sheet("说明")
    workbook.save(path)
    workbook.close()


def test_folder_scan_and_selected_snapshot(tmp_path: Path) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)

    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )

    assert len(catalog.files) == 1
    assert snapshot.name == f"文件夹：{tmp_path.name}"
    assert snapshot.worksheets[0].sourceFile == "scores.xlsx"
    assert snapshot.worksheets[0].sourceSheet == "得分"
    assert snapshot.worksheets[0].dataRows[0] == ["嘟嘟嘟", 33]


def test_folder_catalog_reports_scan_truncation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    for index in range(3):
        _create_source(tmp_path / f"{index}.xlsx")
    monkeypatch.setattr("server.app.folder_workbooks.FILE_LIMIT", 2)

    catalog = scan_folder(tmp_path)

    assert len(catalog.files) == 2
    assert catalog.totalFiles == 3
    assert catalog.truncated is True
    assert catalog.expiresAt


def test_refresh_folder_keeps_session_and_updates_file_list(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)

    renamed = tmp_path / "renamed.xlsx"
    source.rename(renamed)
    (tmp_path / "added.xlsx").write_bytes(b"not an excel file")

    refreshed = refresh_folder(catalog.sessionId)

    assert refreshed.sessionId == catalog.sessionId
    assert refreshed.folderName == catalog.folderName
    names = {file.name for file in refreshed.files}
    assert "renamed.xlsx" in names
    assert "scores.xlsx" not in names
    # 无法读取的文件保留错误信息而不是丢失条目
    added = next(file for file in refreshed.files if file.name == "added.xlsx")
    assert added.error is not None


def test_refresh_folder_reflects_new_worksheets_in_existing_file(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)

    workbook = load_workbook(source)
    workbook.create_sheet("新增")
    workbook.save(source)
    workbook.close()

    refreshed = refresh_folder(catalog.sessionId)

    assert refreshed.files[0].relativePath == "scores.xlsx"
    assert {sheet.name for sheet in refreshed.files[0].worksheets} == {
        "得分",
        "说明",
        "新增",
    }


def test_refresh_folder_rejects_unknown_session() -> None:
    with pytest.raises(ValueError, match="会话已失效"):
        refresh_folder("missing-session")


def test_folder_snapshot_normalizes_dates_codes_and_display_values(
    tmp_path: Path,
) -> None:
    source = tmp_path / "typed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["日期", "编码", "完成率"])
    sheet.append([date(2026, 7, 29), 1, 0.25])
    sheet["B2"].number_format = "000"
    sheet["C2"].number_format = "0%"
    workbook.save(source)
    workbook.close()

    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["数据"])
            ],
        )
    )

    values = snapshot.worksheets[0]
    assert values.dataRows == [["2026-07-29", "001", 0.25]]
    assert values.displayRows == [["2026-07-29", "001", "25%"]]


def test_folder_plan_writes_output_workbook(tmp_path: Path) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    plan = AnalysisPlan.model_validate(
        {
            "id": "folder-output",
            "title": "生成汇总",
            "summary": "写入独立结果文件",
            "actions": [
                {"type": "createWorksheet", "sheet": "汇总"},
                {
                    "type": "writeTable",
                    "sheet": "汇总",
                    "startCell": "A1",
                    "headers": ["人员", "得分"],
                    "rows": [["嘟嘟嘟", 33]],
                },
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(sessionId=catalog.sessionId, plan=plan)
    )

    output = load_workbook(tmp_path / "Excel Bro 结果.xlsx", data_only=True)
    assert result.filesModified == ["Excel Bro 结果.xlsx"]
    assert result.backups == []
    assert len(result.actionResults) == 2
    assert result.verification.status == "verified"
    assert result.verification.passed is True
    assert all(check.passed for check in result.verification.checks)
    assert output["汇总"]["B2"].value == 33
    output.close()


def test_existing_file_is_backed_up_before_write(tmp_path: Path) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    plan = AnalysisPlan.model_validate(
        {
            "id": "folder-edit",
            "title": "修正得分",
            "summary": "修改已有文件",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": "scores.xlsx › 得分",
                    "range": "B2",
                    "values": [[35]],
                }
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    updated = load_workbook(source, data_only=True)
    assert updated["得分"]["B2"].value == 35
    updated.close()
    assert result.filesModified == ["scores.xlsx"]
    assert len(result.backups) == 1
    assert (tmp_path / result.backups[0]).exists()


def test_folder_save_failure_leaves_all_original_files_unchanged(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    original_bytes = source.read_bytes()
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "atomic-save",
            "title": "原子保存",
            "summary": "任一临时文件保存失败时不替换任何目标",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": target,
                    "range": "B2",
                    "values": [[99]],
                },
                {
                    "type": "writeValues",
                    "sheet": "汇总",
                    "range": "A1",
                    "values": [["结果"]],
                },
            ],
        }
    )
    original_save = Workbook.save
    save_count = 0

    def fail_second_save(self, filename) -> None:
        nonlocal save_count
        save_count += 1
        if save_count == 2:
            raise OSError("模拟第二个临时文件保存失败")
        original_save(self, filename)

    monkeypatch.setattr(Workbook, "save", fail_second_save)

    with pytest.raises(OSError, match="第二个临时文件保存失败"):
        execute_folder_plan(
            FolderExecuteRequest(
                sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
            )
        )

    assert source.read_bytes() == original_bytes
    assert not (tmp_path / "Excel Bro 结果.xlsx").exists()
    assert not list(tmp_path.glob("*.excel-bro-backup-*"))
    assert not list(tmp_path.glob(".*.excel-bro-tmp-*"))


def test_folder_execution_closes_workbooks_when_an_action_raises(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    plan = AnalysisPlan.model_validate(
        {
            "id": "close-on-action-error",
            "title": "动作失败释放文件",
            "summary": "图片解析失败时也要关闭工作簿",
            "actions": [
                {
                    "type": "addImage",
                    "sheet": snapshot.worksheets[0].name,
                    "base64": "@@@@",
                    "targetRange": "D1",
                }
            ],
        }
    )
    original_close = Workbook.close
    closed: list[Workbook] = []

    def track_close(self) -> None:
        closed.append(self)
        original_close(self)

    monkeypatch.setattr(Workbook, "close", track_close)

    with pytest.raises(ValueError, match="无法解析图片数据"):
        execute_folder_plan(
            FolderExecuteRequest(
                sessionId=catalog.sessionId,
                plan=_stamp_plan(plan, snapshot),
            )
        )

    assert closed


def test_folder_execution_rejects_files_changed_after_preview(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    changed = load_workbook(source)
    changed["得分"]["B2"] = 99
    changed.save(source)
    changed.close()
    plan = AnalysisPlan.model_validate(
        {
            "id": "stale-preview",
            "title": "拒绝过期预览",
            "summary": "文件变化后不能执行旧计划",
            "sourceFingerprint": snapshot.sourceFingerprint,
            "sourceFingerprintSheets": snapshot.sourceFingerprintSheets,
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": snapshot.worksheets[0].name,
                    "range": "B2",
                    "values": [[35]],
                }
            ],
        }
    )

    with pytest.raises(ValueError, match="发生变化"):
        execute_folder_plan(
            FolderExecuteRequest(sessionId=catalog.sessionId, plan=plan)
        )

    unchanged = load_workbook(source, data_only=True)
    assert unchanged["得分"]["B2"].value == 99
    unchanged.close()


def test_folder_execution_rejects_a_plan_without_preview_fingerprint(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    plan = AnalysisPlan.model_validate(
        {
            "id": "missing-fingerprint",
            "title": "缺少来源指纹",
            "summary": "不能绕过预览新鲜度检查",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": snapshot.worksheets[0].name,
                    "range": "B2",
                    "values": [[99]],
                }
            ],
        }
    )

    with pytest.raises(ValueError, match="缺少数据来源指纹"):
        execute_folder_plan(
            FolderExecuteRequest(sessionId=catalog.sessionId, plan=plan)
        )


def test_plan_cannot_modify_an_unselected_sheet(tmp_path: Path) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    plan = AnalysisPlan.model_validate(
        {
            "id": "unselected-edit",
            "title": "越权修改",
            "summary": "不应执行",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": "scores.xlsx › 说明",
                    "range": "A1",
                    "values": [["不允许"]],
                }
            ],
        }
    )

    with pytest.raises(ValueError, match="未选择"):
        execute_folder_plan(
            FolderExecuteRequest(
                sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
            )
        )


def test_folder_preflight_prevents_earlier_writes_when_later_source_is_missing(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "preflight-missing-source",
            "title": "预检缺失源表",
            "summary": "后续步骤失败时不能先写入前面的值",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": target,
                    "range": "B2",
                    "values": [[99]],
                },
                {
                    "type": "copyRange",
                    "sheet": target,
                    "sourceSheet": "不存在",
                    "sourceRange": "A1",
                    "targetRange": "B3",
                    "copyType": "values",
                    "skipBlanks": False,
                    "transpose": False,
                },
            ],
        }
    )

    with pytest.raises(
        ValueError,
        match="执行前检查未通过.*第 2 步.*复制源工作表",
    ):
        execute_folder_plan(
            FolderExecuteRequest(
                sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
            )
        )

    unchanged = load_workbook(source, data_only=True)
    assert unchanged["得分"]["B2"].value == 33
    unchanged.close()
    assert not list(tmp_path.glob("*.excel-bro-backup-*"))


def test_folder_preflight_rejects_invalid_range_before_writing(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "preflight-invalid-range",
            "title": "预检区域地址",
            "summary": "区域地址错误时不应先写入",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": target,
                    "range": "B2",
                    "values": [[99]],
                },
                {
                    "type": "clearRange",
                    "sheet": target,
                    "range": "not-a-range",
                    "applyTo": "contents",
                },
            ],
        }
    )

    with pytest.raises(
        ValueError,
        match=r"执行前检查未通过.*第 2 步.*区域.*无效",
    ):
        execute_folder_plan(
            FolderExecuteRequest(
                sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
            )
        )

    unchanged = load_workbook(source, data_only=True)
    assert unchanged["得分"]["B2"].value == 33
    unchanged.close()


def test_folder_preflight_rejects_table_and_named_range_conflict(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "preflight-object-name",
            "title": "预检对象名称",
            "summary": "同一计划不能重复使用对象名称",
            "actions": [
                {
                    "type": "createTable",
                    "sheet": target,
                    "range": "A1:B3",
                    "name": "Scores",
                    "hasHeaders": True,
                },
                {
                    "type": "addNamedRange",
                    "sheet": target,
                    "name": "scores",
                    "range": "A1:B3",
                },
            ],
        }
    )

    with pytest.raises(
        ValueError,
        match="执行前检查未通过.*第 2 步.*名称.*已存在",
    ):
        execute_folder_plan(
            FolderExecuteRequest(
                sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
            )
        )

    unchanged = load_workbook(source)
    assert not unchanged["得分"].tables
    unchanged.close()


def test_folder_verification_reports_an_expected_value_mismatch(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    plan = AnalysisPlan.model_validate(
        {
            "id": "failed-verification",
            "title": "验证失败示例",
            "summary": "执行值与验收值不同",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": "结果",
                    "range": "A1",
                    "values": [[1]],
                }
            ],
            "acceptanceCriteria": [
                {
                    "type": "rangeEquals",
                    "sheet": "结果",
                    "range": "A1",
                    "expected": [[2]],
                }
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(sessionId=catalog.sessionId, plan=plan)
    )

    assert result.verification.status == "failed"
    assert result.verification.passed is False
    assert result.verification.checks[0].actual == [[1]]
    assert "与预期不一致" in result.verification.checks[0].message


def test_folder_verifies_sort_filter_table_and_cleared_filter(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    plan = AnalysisPlan.model_validate(
        {
            "id": "verify-structured-actions",
            "title": "排序筛选表格验收",
            "summary": "读取保存后的真实对象状态",
            "actions": [
                {"type": "createWorksheet", "sheet": "汇总"},
                {
                    "type": "writeValues",
                    "sheet": "汇总",
                    "range": "A1:B4",
                    "values": [
                        ["人员", "得分"],
                        ["乙", 20],
                        ["甲", 30],
                        ["丙", None],
                    ],
                },
                {
                    "type": "sortRange",
                    "sheet": "汇总",
                    "range": "A1:B4",
                    "keys": [{"column": 1, "ascending": False}],
                    "hasHeaders": True,
                },
                {
                    "type": "filterRange",
                    "sheet": "汇总",
                    "range": "A1:B4",
                    "column": 0,
                    "values": ["甲", "乙"],
                },
                {
                    "type": "createTable",
                    "sheet": "汇总",
                    "range": "A1:B4",
                    "name": "ResultTable",
                    "hasHeaders": True,
                },
            ],
            "acceptanceCriteria": [
                {
                    "type": "rangeEquals",
                    "sheet": "汇总",
                    "range": "A1:B4",
                    "expected": [
                        ["人员", "得分"],
                        ["甲", 30],
                        ["乙", 20],
                        ["丙", None],
                    ],
                }
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(sessionId=catalog.sessionId, plan=plan)
    )

    assert result.verification.status == "verified"
    assert result.verification.passed is True
    assert {check.criterion.type for check in result.verification.checks} >= {
        "rangeSorted",
        "filterApplied",
        "tableExists",
    }
    assert all(check.passed for check in result.verification.checks)

    clear_plan = AnalysisPlan.model_validate(
        {
            "id": "verify-cleared-filter",
            "title": "清除筛选验收",
            "summary": "确认筛选条件已清除",
            "actions": [{"type": "clearFilter", "sheet": "汇总"}],
        }
    )
    cleared = execute_folder_plan(
        FolderExecuteRequest(sessionId=catalog.sessionId, plan=clear_plan)
    )
    assert cleared.verification.status == "verified"
    assert cleared.verification.checks[-1].criterion.type == "filterCleared"
    assert cleared.verification.checks[-1].passed is True


def test_folder_mode_writes_formulas_formats_and_clears_ranges(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "advanced-edit",
            "title": "公式与格式",
            "summary": "写入公式、设置格式并清空旧值",
            "actions": [
                {
                    "type": "writeFormulas",
                    "sheet": target,
                    "range": "C2",
                    "formulas": [["=SUM(B2:B4)"]],
                },
                {
                    "type": "setNumberFormat",
                    "sheet": target,
                    "range": "C2",
                    "formatCode": "0.00",
                },
                {
                    "type": "clearRange",
                    "sheet": target,
                    "range": "B2",
                    "applyTo": "contents",
                },
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    updated = load_workbook(source, data_only=False)
    assert updated["得分"]["C2"].value == "=SUM(B2:B4)"
    assert updated["得分"]["C2"].number_format == "0.00"
    assert updated["得分"]["B2"].value is None
    updated.close()
    assert result.verification.status == "verified"
    assert result.verification.passed is True
    assert result.verification.unverifiedActions == []


def test_folder_mode_quotes_formula_like_values_to_prevent_injection(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "inject-guard",
            "title": "防公式注入",
            "summary": "值路径的 = 开头字符串应按文本存储",
            "actions": [
                {
                    "type": "writeValues",
                    "sheet": target,
                    "range": "B2",
                    "values": [[r'=WEBSERVICE("http://evil.com")']],
                },
                {
                    "type": "writeTable",
                    "sheet": target,
                    "startCell": "D1",
                    "headers": [r"=EXEC('calc')"],
                    "rows": [[r"=cmd|'/c calc'!A0"]],
                },
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    updated = load_workbook(source, data_only=False)
    ws = updated["得分"]
    b2 = ws["B2"]
    d1 = ws["D1"]
    d2 = ws["D2"]
    # 值路径注入被强制为文本：data_type 是字符串而非公式，且带前导撇号。
    assert b2.data_type == "s"
    assert b2.value == '\'=WEBSERVICE("http://evil.com")'
    assert d1.data_type == "s"
    assert d1.value == "'=EXEC('calc')"
    assert d2.data_type == "s"
    assert d2.value == "'=cmd|'/c calc'!A0"
    updated.close()
    assert result.verification.status == "verified"
    assert result.verification.passed is True
    assert result.verification.unverifiedActions == []


def test_folder_verifies_formats_validation_and_freeze_panes(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "verify-formatting",
            "title": "格式与验证验收",
            "summary": "检查保存后的真实格式、验证规则和冻结位置",
            "actions": [
                {
                    "type": "setFill",
                    "sheet": target,
                    "range": "A2:B3",
                    "color": "#DFF3E4",
                },
                {
                    "type": "setFont",
                    "sheet": target,
                    "range": "A2:B3",
                    "bold": True,
                    "color": "#123456",
                },
                {
                    "type": "setNumberFormat",
                    "sheet": target,
                    "range": "B2:B3",
                    "formatCode": "0.00",
                },
                {
                    "type": "setBorders",
                    "sheet": target,
                    "range": "A2:B3",
                    "sides": [
                        "top",
                        "bottom",
                        "left",
                        "right",
                        "insideHorizontal",
                        "insideVertical",
                    ],
                    "style": "continuous",
                    "color": "#445566",
                    "weight": "medium",
                },
                {
                    "type": "setAlignment",
                    "sheet": target,
                    "range": "A2:B3",
                    "horizontal": "center",
                    "vertical": "bottom",
                    "wrapText": True,
                },
                {
                    "type": "resizeRange",
                    "sheet": target,
                    "range": "A2:B3",
                    "rowHeight": 24,
                    "columnWidth": 18,
                },
                {
                    "type": "setDataValidation",
                    "sheet": target,
                    "range": "B2:B3",
                    "validationType": "wholeNumber",
                    "formula1": 0,
                    "formula2": 100,
                    "operator": "between",
                    "allowBlank": False,
                    "prompt": "请输入 0 到 100",
                    "errorMessage": "分数超出范围",
                },
                {
                    "type": "freezePanes",
                    "sheet": target,
                    "rows": 1,
                    "columns": 1,
                },
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    assert result.verification.status == "verified"
    assert result.verification.passed is True
    assert result.verification.unverifiedActions == []
    assert all(check.passed for check in result.verification.checks)
    assert {check.criterion.type for check in result.verification.checks} >= {
        "rangeFormatMatches",
        "bordersMatch",
        "dataValidationMatches",
        "freezePanesMatches",
    }


def test_folder_keeps_unsupported_format_verification_explicit(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["得分"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "unverified-conditional-format",
            "title": "条件格式弱验收",
            "summary": "无法稳定读取的属性保持未独立验证",
            "actions": [
                {
                    "type": "setConditionalFormat",
                    "sheet": target,
                    "range": "B2:B3",
                    "ruleType": "cellValue",
                    "operator": "greaterThan",
                    "formula1": 40,
                    "color": "#FFF2CC",
                }
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    assert result.verification.status == "executed_unverified"
    assert [gap.type for gap in result.verification.unverifiedActions] == [
        "setConditionalFormat"
    ]


def test_folder_mode_deletes_a_selected_sheet_and_verifies_it(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scores.xlsx"
    _create_source(source)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["说明"])
            ],
        )
    )
    target = snapshot.worksheets[0].name
    plan = AnalysisPlan.model_validate(
        {
            "id": "delete-sheet",
            "title": "删除说明",
            "summary": "删除已选说明工作表",
            "actions": [{"type": "deleteWorksheet", "sheet": target}],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    updated = load_workbook(source)
    assert "说明" not in updated.sheetnames
    updated.close()
    assert result.verification.status == "verified"
    assert result.verification.passed is True


def test_folder_mode_splits_full_table_and_calculates_group_ratios(
    tmp_path: Path,
) -> None:
    source = tmp_path / "cinema.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet-dc1"
    sheet.append(["hunan(制表日期:2026-07-25)", None, None, None])
    sheet.append([None, None, None, None])
    sheet.append(["影院名称", "影院编码", "影厅", "影片名称"])
    sheet.append(["怀化横店电影城", "43122301", "1号厅", "恐怖游轮"])
    sheet.append(["怀化横店电影城", "43122301", "2号厅", "恐怖游轮"])
    sheet.append(["怀化横店电影城", "43122301", "3号厅", "功夫女足"])
    sheet.append(["广州影院", "44010001", "1号厅", "恐怖游轮"])
    sheet.append(["长沙影院", "43010001", "1号厅", "三国第一部：争洛阳"])
    workbook.save(source)
    workbook.close()

    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(
                    fileId=catalog.files[0].id, sheets=["sheet-dc1"]
                )
            ],
        )
    )
    plan = AnalysisPlan.model_validate(
        {
            "id": "split-cinema",
            "title": "按影片拆分",
            "summary": "按影院统计影片记录数和占比",
            "actions": [
                {
                    "type": "splitGroupAggregate",
                    "sheet": snapshot.worksheets[0].name,
                    "splitBy": "影片名称",
                    "groupBy": ["影院名称", "影院编码"],
                    "metrics": [
                        {
                            "operation": "countRows",
                            "outputName": "计数项:影厅",
                            "ratioOutputName": "占比",
                        }
                    ],
                }
            ],
        }
    )

    result = execute_folder_plan(
        FolderExecuteRequest(
            sessionId=catalog.sessionId, plan=_stamp_plan(plan, snapshot)
        )
    )

    updated = load_workbook(source, data_only=True)
    assert "恐怖游轮" in updated.sheetnames
    assert "功夫女足" in updated.sheetnames
    assert "三国第一部 争洛阳" in updated.sheetnames
    horror = updated["恐怖游轮"]
    assert [cell.value for cell in horror[1]] == [
        "影院名称",
        "影院编码",
        "影片名称",
        "计数项:影厅",
        "占比",
    ]
    assert [cell.value for cell in horror[2]] == [
        "怀化横店电影城",
        "43122301",
        "恐怖游轮",
        2,
        pytest.approx(2 / 3),
    ]
    assert horror["E2"].number_format == "0.00%"
    updated.close()
    assert result.verification.status == "executed_unverified"
    assert result.verification.passed is False
    assert result.verification.unverifiedActions[0].type == "splitGroupAggregate"
