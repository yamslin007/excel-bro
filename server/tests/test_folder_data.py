from pathlib import Path

from openpyxl import Workbook

from server.app.folder_workbooks import (
    FolderQueryRequest,
    FolderSelection,
    FolderSnapshotRequest,
    create_folder_snapshot,
    query_folder_data,
    scan_folder,
)
from server.app.models import DataToolRequest


def _create_rows(path: Path, start: int, count: int) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["分类", "数值"])
    for value in range(start, start + count):
        sheet.append(["甲" if value % 2 else "乙", value])
    workbook.save(path)
    workbook.close()


def test_folder_query_reads_all_rows_beyond_snapshot_limit(tmp_path: Path) -> None:
    source = tmp_path / "large.xlsx"
    _create_rows(source, 1, 300)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["数据"])
            ],
        )
    )
    assert snapshot.worksheets[0].truncated is True

    result = query_folder_data(
        FolderQueryRequest(
            sessionId=catalog.sessionId,
            request=DataToolRequest.model_validate(
                {
                    "id": "full-data",
                    "tool": "query_table",
                    "arguments": {
                        "mode": "aggregate",
                        "metrics": [
                            {
                                "operation": "sum",
                                "field": "数值",
                                "outputName": "合计",
                            },
                            {
                                "operation": "average",
                                "field": "数值",
                                "outputName": "平均值",
                            },
                        ],
                    },
                }
            ),
        )
    )

    assert result.scannedRows == 300
    assert result.rows == [[45150, 150.5]]


def test_same_sheet_names_keep_distinct_file_and_sheet_ids(tmp_path: Path) -> None:
    _create_rows(tmp_path / "a.xlsx", 1, 2)
    _create_rows(tmp_path / "b.xlsx", 10, 2)
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=file.id, sheets=["数据"])
                for file in catalog.files
            ],
        )
    )

    assert len({sheet.sourceFileId for sheet in snapshot.worksheets}) == 2
    assert len({sheet.sourceSheetId for sheet in snapshot.worksheets}) == 2
    result = query_folder_data(
        FolderQueryRequest(
            sessionId=catalog.sessionId,
            request=DataToolRequest.model_validate(
                {
                    "id": "two-files",
                    "arguments": {
                        "mode": "rows",
                        "fields": ["来源文件ID", "来源工作表ID", "数值"],
                        "limit": 10,
                    },
                }
            ),
        )
    )
    assert result.scannedRows == 4
    assert len({row[0] for row in result.rows}) == 2
    assert len({row[1] for row in result.rows}) == 2


def test_unselected_file_is_not_read_by_folder_query(tmp_path: Path) -> None:
    _create_rows(tmp_path / "selected.xlsx", 1, 2)
    _create_rows(tmp_path / "unselected.xlsx", 100, 2)
    catalog = scan_folder(tmp_path)
    selected = next(file for file in catalog.files if file.name == "selected.xlsx")
    create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[FolderSelection(fileId=selected.id, sheets=["数据"])],
        )
    )

    result = query_folder_data(
        FolderQueryRequest(
            sessionId=catalog.sessionId,
            request=DataToolRequest.model_validate(
                {
                    "id": "authorized-only",
                    "arguments": {
                        "mode": "aggregate",
                        "metrics": [
                            {
                                "operation": "sum",
                                "field": "数值",
                                "outputName": "合计",
                            }
                        ],
                    },
                }
            ),
        )
    )

    assert result.scannedRows == 2
    assert result.rows == [[3]]
    assert result.sourceSheets == ["selected.xlsx › 数据"]


def test_folder_query_supports_deduplicate_and_join(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["编号", "姓名"])
    sheet.append([1, "甲"])
    sheet.append([2, "乙"])
    workbook.save(first)
    workbook.close()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["编号", "金额"])
    sheet.append([2, 20])
    sheet.append([3, 30])
    workbook.save(second)
    workbook.close()
    catalog = scan_folder(tmp_path)
    snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=file.id, sheets=["数据"])
                for file in catalog.files
            ],
        )
    )

    deduplicated = query_folder_data(
        FolderQueryRequest(
            sessionId=catalog.sessionId,
            request=DataToolRequest.model_validate(
                {
                    "id": "deduplicate",
                    "arguments": {
                        "mode": "rows",
                        "fields": ["编号"],
                        "combine": {
                            "mode": "deduplicate",
                            "deduplicateBy": ["编号"],
                        },
                        "sortBy": "编号",
                        "sortDirection": "asc",
                    },
                }
            ),
        )
    )
    assert deduplicated.rows == [[1], [2], [3]]

    left, right = snapshot.worksheets
    joined = query_folder_data(
        FolderQueryRequest(
            sessionId=catalog.sessionId,
            request=DataToolRequest.model_validate(
                {
                    "id": "join",
                    "arguments": {
                        "mode": "rows",
                        "fields": ["编号", "姓名", "金额"],
                        "combine": {
                            "mode": "join",
                            "leftSourceSheetId": left.sourceSheetId,
                            "rightSourceSheetId": right.sourceSheetId,
                            "leftKey": "编号",
                            "rightKey": "编号",
                            "joinHow": "inner",
                        },
                    },
                }
            ),
        )
    )
    assert joined.rows == [[2, "乙", 20]]


def test_folder_query_preserves_formatted_identifier_text(tmp_path: Path) -> None:
    source = tmp_path / "codes.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["编码", "数值"])
    sheet.append([1, 10])
    sheet["A2"].number_format = "000"
    workbook.save(source)
    workbook.close()
    catalog = scan_folder(tmp_path)
    create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=catalog.sessionId,
            selections=[
                FolderSelection(fileId=catalog.files[0].id, sheets=["数据"])
            ],
        )
    )

    result = query_folder_data(
        FolderQueryRequest(
            sessionId=catalog.sessionId,
            request=DataToolRequest.model_validate(
                {
                    "id": "formatted-code",
                    "arguments": {
                        "mode": "rows",
                        "fields": ["编码"],
                    },
                }
            ),
        )
    )

    assert result.rows == [["001"]]


def test_folder_and_sheet_ids_are_stable_across_rescans(tmp_path: Path) -> None:
    _create_rows(tmp_path / "stable.xlsx", 1, 1)
    first = scan_folder(tmp_path)
    first_snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=first.sessionId,
            selections=[
                FolderSelection(fileId=first.files[0].id, sheets=["数据"])
            ],
        )
    )
    second = scan_folder(tmp_path)
    second_snapshot = create_folder_snapshot(
        FolderSnapshotRequest(
            sessionId=second.sessionId,
            selections=[
                FolderSelection(fileId=second.files[0].id, sheets=["数据"])
            ],
        )
    )

    assert first.files[0].id == second.files[0].id
    assert (
        first_snapshot.worksheets[0].sourceSheetId
        == second_snapshot.worksheets[0].sourceSheetId
    )
