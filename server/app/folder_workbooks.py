from __future__ import annotations

import base64
import hashlib
import io
import re
import shutil
import uuid
import time
from contextvars import ContextVar
from copy import copy
from datetime import date, datetime
from functools import cmp_to_key
from pathlib import Path
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.chart import AreaChart, BarChart, DoughnutChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import FilterColumn, Filters
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from pydantic import BaseModel, Field

from .models import (
    ActionExecutionResult,
    AnalysisPlan,
    DataToolRequest,
    DataToolResult,
    UnverifiedAction,
    VerificationCheck,
    VerificationReport,
    WorkbookSnapshot,
    WorksheetSnapshot,
)
from .capabilities import capability_int, capability_text

ROW_LIMIT = capability_int("snapshot", "dataRows")
COLUMN_LIMIT = capability_int("snapshot", "dataColumns")
FILE_LIMIT = capability_int("folder", "maxFiles")
SOURCE_SEPARATOR = " › "
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_OPEN_WRITE_WORKBOOKS: ContextVar[list[Workbook] | None] = ContextVar(
    "open_write_workbooks", default=None
)
OUTPUT_FILE_NAME = capability_text("folder", "outputFileName")


class FolderWorksheetInfo(BaseModel):
    name: str
    rowCount: int
    columnCount: int


class FolderFileInfo(BaseModel):
    id: str
    name: str
    relativePath: str
    worksheets: list[FolderWorksheetInfo] = Field(default_factory=list)
    error: str | None = None


class FolderCatalog(BaseModel):
    sessionId: str
    folderName: str
    folderPath: str
    files: list[FolderFileInfo]
    totalFiles: int
    truncated: bool
    expiresAt: str


class FolderSelection(BaseModel):
    fileId: str
    sheets: list[str] = Field(min_length=1)


class FolderSnapshotRequest(BaseModel):
    sessionId: str
    selections: list[FolderSelection] = Field(min_length=1)


class FolderExecuteRequest(BaseModel):
    sessionId: str
    plan: AnalysisPlan


class FolderExecuteResponse(BaseModel):
    filesModified: list[str]
    backups: list[str]
    actionResults: list[ActionExecutionResult]
    verification: VerificationReport
    executionMs: float = Field(ge=0)
    verificationMs: float = Field(ge=0)


class FolderQueryRequest(BaseModel):
    sessionId: str
    request: DataToolRequest


class _FolderSession:
    def __init__(self, root: Path, files: dict[str, Path]) -> None:
        self.root = root
        self.files = files
        self.selected_targets: dict[str, tuple[Path, str]] = {}
        self.selected_sources: dict[str, tuple[str, str, Path, str]] = {}
        self.source_fingerprint: str | None = None
        self.last_access = time.monotonic()


_sessions: dict[str, _FolderSession] = {}


def _prune_sessions(*, reserve_slot: bool = False) -> None:
    ttl = capability_int("folder", "sessionTtlSeconds")
    cutoff = time.monotonic() - ttl
    expired = [
        session_id
        for session_id, session in _sessions.items()
        if session.last_access < cutoff
    ]
    for session_id in expired:
        _sessions.pop(session_id, None)
    capacity = capability_int("folder", "maxSessions")
    target_size = capacity - 1 if reserve_slot else capacity
    while len(_sessions) > target_size:
        oldest = min(_sessions, key=lambda key: _sessions[key].last_access)
        _sessions.pop(oldest, None)


def choose_folder() -> Path | None:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.askdirectory(title="选择包含 Excel 文件的文件夹")
    finally:
        root.destroy()
    return Path(selected).resolve() if selected else None


def _is_supported_file(path: Path) -> bool:
    return (
        path.is_file()
        and path.suffix.lower() in SUPPORTED_SUFFIXES
        and not path.name.startswith("~$")
        and ".excel-bro-backup-" not in path.name
    )


def scan_folder(root: Path) -> FolderCatalog:
    root = root.resolve()
    all_candidates = sorted(
        (path for path in root.rglob("*") if _is_supported_file(path)),
        key=lambda path: str(path.relative_to(root)).casefold(),
    )
    candidates = all_candidates[:FILE_LIMIT]
    _prune_sessions(reserve_slot=True)
    session_id = uuid.uuid4().hex
    session_files: dict[str, Path] = {}
    files: list[FolderFileInfo] = []

    for path in candidates:
        file_id = uuid.uuid5(
            uuid.NAMESPACE_URL,
            str(path.relative_to(root)).replace("\\", "/").casefold(),
        ).hex
        session_files[file_id] = path
        relative_path = str(path.relative_to(root))
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_vba=path.suffix.lower() == ".xlsm",
            )
            worksheets = [
                FolderWorksheetInfo(
                    name=sheet.title,
                    rowCount=sheet.max_row,
                    columnCount=sheet.max_column,
                )
                for sheet in workbook.worksheets
            ]
            workbook.close()
            files.append(
                FolderFileInfo(
                    id=file_id,
                    name=path.name,
                    relativePath=relative_path,
                    worksheets=worksheets,
                )
            )
        except Exception as error:
            files.append(
                FolderFileInfo(
                    id=file_id,
                    name=path.name,
                    relativePath=relative_path,
                    error=f"无法读取：{error}",
                )
            )

    _sessions[session_id] = _FolderSession(root, session_files)
    return FolderCatalog(
        sessionId=session_id,
        folderName=root.name,
        folderPath=str(root),
        files=files,
        totalFiles=len(all_candidates),
        truncated=len(all_candidates) > FILE_LIMIT,
        expiresAt=datetime.fromtimestamp(
            datetime.now().timestamp()
            + capability_int("folder", "sessionTtlSeconds")
        ).astimezone().isoformat(),
    )


def select_and_scan_folder() -> FolderCatalog | None:
    selected = choose_folder()
    return scan_folder(selected) if selected else None


def _session(session_id: str) -> _FolderSession:
    _prune_sessions()
    try:
        session = _sessions[session_id]
        session.last_access = time.monotonic()
        return session
    except KeyError as error:
        raise ValueError("文件夹会话已失效，请重新选择文件夹") from error


def _selected_files_fingerprint(paths: set[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths, key=lambda item: str(item).casefold()):
        digest.update(str(path).encode("utf-8"))
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
    return digest.hexdigest()


def _normalized_openpyxl_value(cell):
    value = cell.value
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        if (
            isinstance(value, (int, float))
            and re.fullmatch(r"0+", cell.number_format or "")
        ):
            return f"{int(value):0{len(cell.number_format)}d}"
        return value
    return str(value)


def _display_openpyxl_value(cell):
    value = _normalized_openpyxl_value(cell)
    if (
        isinstance(cell.value, (int, float))
        and "%" in (cell.number_format or "")
    ):
        decimals = max(0, (cell.number_format.split("%", 1)[0].partition(".")[2].count("0")))
        return f"{cell.value * 100:.{decimals}f}%"
    return value


def create_folder_snapshot(request: FolderSnapshotRequest) -> WorkbookSnapshot:
    session = _session(request.sessionId)
    snapshots: list[WorksheetSnapshot] = []
    selected_targets: dict[str, tuple[Path, str]] = {}
    selected_sources: dict[str, tuple[str, str, Path, str]] = {}
    active_name = ""

    for selection in request.selections:
        try:
            path = session.files[selection.fileId]
        except KeyError as error:
            raise ValueError("选择中包含未知文件，请重新扫描文件夹") from error
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_vba=path.suffix.lower() == ".xlsm",
        )
        available = set(workbook.sheetnames)
        relative_path = str(path.relative_to(session.root))
        for sheet_name in selection.sheets:
            if sheet_name not in available:
                workbook.close()
                raise ValueError(f"{relative_path} 中不存在工作表「{sheet_name}」")
            sheet = workbook[sheet_name]
            row_count = sheet.max_row
            column_count = sheet.max_column
            cells = [
                list(row)
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(row_count, ROW_LIMIT + 1),
                    max_col=min(column_count, COLUMN_LIMIT),
                )
            ]
            values = [
                [_normalized_openpyxl_value(cell) for cell in row]
                for row in cells
            ]
            display_values = [
                [_display_openpyxl_value(cell) for cell in row]
                for row in cells
            ]
            display_name = f"{relative_path}{SOURCE_SEPARATOR}{sheet_name}"
            sheet_id = uuid.uuid5(
                uuid.NAMESPACE_URL,
                f"{selection.fileId}:{sheet_name}",
            ).hex
            selected_targets[display_name] = (path, sheet_name)
            selected_sources[display_name] = (
                selection.fileId,
                sheet_id,
                path,
                sheet_name,
            )
            snapshots.append(
                WorksheetSnapshot(
                    name=display_name,
                    sourceFileId=selection.fileId,
                    sourceSheetId=sheet_id,
                    sourceFile=relative_path,
                    sourceSheet=sheet_name,
                    usedRange=(
                        f"A1:{get_column_letter(column_count)}{row_count}"
                        if row_count and column_count
                        else None
                    ),
                    rowCount=row_count,
                    columnCount=column_count,
                    headers=list(values[0]) if values else [],
                    dataRows=[list(row) for row in values[1:]],
                    displayRows=[list(row) for row in display_values[1:]],
                    truncated=row_count > ROW_LIMIT + 1
                    or column_count > COLUMN_LIMIT,
                )
            )
            if not active_name:
                active_name = display_name
        workbook.close()

    if not snapshots:
        raise ValueError("请至少选择一个可读取的工作表")
    session.selected_targets = selected_targets
    session.selected_sources = selected_sources
    session.source_fingerprint = _selected_files_fingerprint(
        {path for path, _ in selected_targets.values()}
    )
    return WorkbookSnapshot(
        name=f"文件夹：{session.root.name}",
        capturedAt=datetime.now().astimezone().isoformat(),
        activeWorksheet=active_name,
        sourceFingerprint=session.source_fingerprint,
        sourceFingerprintSheets=list(selected_targets),
        worksheets=snapshots,
    )


def query_folder_data(request: FolderQueryRequest) -> DataToolResult:
    from .folder_data import AuthorizedSheet, execute_folder_query

    session = _session(request.sessionId)
    if not session.selected_sources:
        raise ValueError("文件夹数据范围尚未确认，请先生成预览")
    sources = [
        AuthorizedSheet(
            file_id=file_id,
            sheet_id=sheet_id,
            path=path,
            sheet_name=sheet_name,
            display_name=display_name,
        )
        for display_name, (
            file_id,
            sheet_id,
            path,
            sheet_name,
        ) in session.selected_sources.items()
    ]
    return execute_folder_query(request.request, sources)


def _target(
    session: _FolderSession, sheet_reference: str
) -> tuple[Path, str, Literal["source", "output"]]:
    if sheet_reference in session.selected_targets:
        path, sheet_name = session.selected_targets[sheet_reference]
        return path, sheet_name, "source"
    if SOURCE_SEPARATOR in sheet_reference:
        raise ValueError(f"计划试图操作未选择的工作表：{sheet_reference}")
    return session.root / OUTPUT_FILE_NAME, sheet_reference, "output"


def _load_for_write(path: Path) -> Workbook:
    if path.exists():
        workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    else:
        workbook = Workbook()
    active = _OPEN_WRITE_WORKBOOKS.get()
    if active is not None:
        active.append(workbook)
    return workbook


def _get_or_create_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    if (
        len(workbook.sheetnames) == 1
        and workbook.active.title == "Sheet"
        and workbook.active.max_row == 1
        and workbook.active.max_column == 1
        and workbook.active["A1"].value is None
    ):
        workbook.active.title = sheet_name
        return workbook.active
    return workbook.create_sheet(sheet_name)


def _values_equal(actual: object, expected: object) -> bool:
    if actual in (None, "") and expected in (None, ""):
        return True
    if (
        isinstance(actual, (int, float))
        and not isinstance(actual, bool)
        and isinstance(expected, (int, float))
        and not isinstance(expected, bool)
    ):
        return float(actual) == float(expected)
    return actual == expected


def _normalized_range(address: str) -> str:
    local_address = address.rsplit("!", 1)[-1]
    return local_address.replace("$", "").strip().upper()


def _compare_sort_values(left: object, right: object) -> int:
    left_blank = left in (None, "")
    right_blank = right in (None, "")
    if left_blank or right_blank:
        if left_blank and right_blank:
            return 0
        return 1 if left_blank else -1
    if (
        isinstance(left, (int, float))
        and not isinstance(left, bool)
        and isinstance(right, (int, float))
        and not isinstance(right, bool)
    ):
        return (float(left) > float(right)) - (float(left) < float(right))
    if isinstance(left, bool) and isinstance(right, bool):
        return int(left) - int(right)
    left_text = str(left).casefold()
    right_text = str(right).casefold()
    return (left_text > right_text) - (left_text < right_text)


def _compare_sorted_rows(left: list[object], right: list[object], keys) -> int:
    for key in keys:
        left_value = left[key.column]
        right_value = right[key.column]
        comparison = _compare_sort_values(left_value, right_value)
        if comparison == 0:
            continue
        if left_value in (None, "") or right_value in (None, ""):
            return comparison
        return comparison if key.ascending else -comparison
    return 0


def _range_is_sorted(values: list[list[object]], keys, has_headers: bool) -> bool:
    rows = values[1:] if has_headers else values
    if not keys or any(
        key.column < 0 or any(key.column >= len(row) for row in rows)
        for key in keys
    ):
        return False
    return all(
        _compare_sorted_rows(rows[index - 1], rows[index], keys) <= 0
        for index in range(1, len(rows))
    )


def _normalized_color(value: object) -> str:
    rgb = getattr(value, "rgb", value)
    text = str(rgb or "").removeprefix("#")
    return f"#{text[-6:].upper()}" if len(text) >= 6 else text.upper()


def _openpyxl_border_style(style: str, weight: str) -> str | None:
    if style == "none":
        return None
    if style == "continuous":
        return {
            "hairline": "hair",
            "thin": "thin",
            "medium": "medium",
            "thick": "thick",
        }[weight]
    if style == "dash":
        return "mediumDashed" if weight in {"medium", "thick"} else "dashed"
    if style == "dashDot":
        return "mediumDashDot" if weight in {"medium", "thick"} else "dashDot"
    if style == "dot":
        return "dotted"
    return "double"


def _range_format_matches(sheet, criterion) -> bool:
    cells = [cell for row in _iter_range(sheet, criterion.range) for cell in row]
    if criterion.fillColor is not None and not all(
        cell.fill.fill_type == "solid"
        and _normalized_color(cell.fill.fgColor) == _normalized_color(criterion.fillColor)
        for cell in cells
    ):
        return False
    if criterion.bold is not None and not all(
        cell.font.bold == criterion.bold for cell in cells
    ):
        return False
    if criterion.fontColor is not None and not all(
        _normalized_color(cell.font.color) == _normalized_color(criterion.fontColor)
        for cell in cells
    ):
        return False
    if criterion.numberFormat is not None and not all(
        cell.number_format == criterion.numberFormat for cell in cells
    ):
        return False
    expected_horizontal = {
        "centerAcrossSelection": "centerContinuous"
    }.get(criterion.horizontal, criterion.horizontal)
    if criterion.horizontal is not None and not all(
        cell.alignment.horizontal == expected_horizontal for cell in cells
    ):
        return False
    if criterion.vertical is not None and not all(
        cell.alignment.vertical == criterion.vertical for cell in cells
    ):
        return False
    if criterion.wrapText is not None and not all(
        cell.alignment.wrap_text == criterion.wrapText for cell in cells
    ):
        return False
    min_col, min_row, max_col, max_row = range_boundaries(criterion.range)
    if criterion.rowHeight is not None and not all(
        sheet.row_dimensions[row].height is not None
        and abs(sheet.row_dimensions[row].height - criterion.rowHeight) <= 0.1
        for row in range(min_row, max_row + 1)
    ):
        return False
    if criterion.columnWidth is not None and not all(
        sheet.column_dimensions[get_column_letter(column)].width is not None
        and abs(
            sheet.column_dimensions[get_column_letter(column)].width
            - criterion.columnWidth
        )
        <= 0.1
        for column in range(min_col, max_col + 1)
    ):
        return False
    return True


def _borders_match(sheet, criterion) -> bool:
    side_names = {
        "top": "top",
        "bottom": "bottom",
        "left": "left",
        "right": "right",
        "insideHorizontal": "horizontal",
        "insideVertical": "vertical",
    }
    expected_style = _openpyxl_border_style(criterion.style, criterion.weight)
    for row in _iter_range(sheet, criterion.range):
        for cell in row:
            for side in criterion.sides:
                actual = getattr(cell.border, side_names[side])
                if actual.style != expected_style:
                    return False
                if (
                    expected_style is not None
                    and _normalized_color(actual.color)
                    != _normalized_color(criterion.color)
                ):
                    return False
    return True


def _data_validation_matches(sheet, criterion) -> bool:
    expected_type = {
        "wholeNumber": "whole",
        "textLength": "textLength",
    }.get(criterion.validationType, criterion.validationType)
    expected_operator = {
        "equalTo": "equal",
        "notEqualTo": "notEqual",
        "greaterThanOrEqualTo": "greaterThanOrEqual",
        "lessThanOrEqualTo": "lessThanOrEqual",
    }.get(criterion.operator, criterion.operator)
    expected_formula1 = (
        '"' + ",".join(str(value) for value in criterion.values) + '"'
        if expected_type == "list"
        else criterion.formula1
    )
    for validation in sheet.data_validations.dataValidation:
        ranges = {
            _normalized_range(str(cell_range))
            for cell_range in validation.sqref.ranges
        }
        if _normalized_range(criterion.range) not in ranges:
            continue
        return (
            validation.type == expected_type
            and validation.operator == expected_operator
            and str(
                "" if validation.formula1 is None else validation.formula1
            )
            == str("" if expected_formula1 is None else expected_formula1)
            and str(
                "" if validation.formula2 is None else validation.formula2
            )
            == str("" if criterion.formula2 is None else criterion.formula2)
            and bool(validation.allow_blank) == criterion.allowBlank
            and (
                criterion.prompt is None
                or validation.prompt == criterion.prompt
            )
            and (
                criterion.errorMessage is None
                or validation.error == criterion.errorMessage
            )
        )
    return False


def _freeze_panes_position(sheet) -> tuple[int, int]:
    location = sheet.freeze_panes
    if location is None:
        return 0, 0
    coordinate = getattr(location, "coordinate", str(location))
    min_col, min_row, _, _ = range_boundaries(coordinate)
    return min_row - 1, min_col - 1


def _iter_range(sheet, address: str):
    min_column, min_row, max_column, max_row = range_boundaries(address)
    return sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    )


def _write_matrix(sheet, address: str, values: list[list[object]]) -> None:
    min_column, min_row, max_column, max_row = range_boundaries(address)
    if (
        max_row - min_row + 1 != len(values)
        or not values
        or max_column - min_column + 1 != len(values[0])
    ):
        raise ValueError(f"{address} 与写入数据尺寸不一致")
    for row_offset, row in enumerate(values):
        for column_offset, value in enumerate(row):
            sheet.cell(
                row=min_row + row_offset,
                column=min_column + column_offset,
                value=value,
            )


def _folder_mode_unsupported(action_type: str) -> None:
    raise ValueError(
        f"文件夹模式暂不支持 {action_type}；请打开工作簿后在 Excel 内执行该动作"
    )


def _preflight_range_shape(address: str) -> tuple[int, int]:
    min_column, min_row, max_column, max_row = range_boundaries(address)
    if min_column < 1 or min_row < 1 or max_column > 16_384 or max_row > 1_048_576:
        raise ValueError("超出 Excel 工作表边界")
    return max_row - min_row + 1, max_column - min_column + 1


def _action_range_fields(action) -> list[tuple[str, str]]:
    fields: list[tuple[str, str]] = []
    for attribute, label in (
        ("range", "区域"),
        ("startCell", "起始位置"),
        ("cell", "单元格"),
        ("sourceRange", "源区域"),
        ("targetRange", "目标区域"),
        ("destinationCell", "目标位置"),
    ):
        value = getattr(action, attribute, None)
        if value:
            fields.append((label, value))
    return fields


def _preflight_folder_plan(
    session: _FolderSession,
    plan: AnalysisPlan,
) -> None:
    unsupported = {
        "insertRange",
        "deleteRange",
        "createPivotTable",
        "addShape",
    }
    sheet_names: dict[Path, set[str]] = {}
    object_names: dict[Path, set[str]] = {}

    def names_for(path: Path) -> set[str]:
        existing = sheet_names.get(path)
        if existing is not None:
            return existing
        if path.exists():
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_vba=path.suffix.lower() == ".xlsm",
            )
            try:
                existing = set(workbook.sheetnames)
            finally:
                workbook.close()
        else:
            existing = {"Sheet"}
        sheet_names[path] = existing
        return existing

    def objects_for(path: Path) -> set[str]:
        existing = object_names.get(path)
        if existing is not None:
            return existing
        if not path.exists():
            existing = set()
        else:
            workbook = load_workbook(
                path,
                read_only=False,
                data_only=False,
                keep_vba=path.suffix.lower() == ".xlsm",
            )
            try:
                existing = {
                    name.casefold() for name in workbook.defined_names.keys()
                }
                existing.update(
                    table.name.casefold()
                    for worksheet in workbook.worksheets
                    for table in worksheet.tables.values()
                )
            finally:
                workbook.close()
        object_names[path] = existing
        return existing

    issues: list[str] = []
    for index, action in enumerate(plan.actions):
        if action.type in unsupported:
            issues.append(
                f"第 {index + 1} 步：文件夹模式暂不支持 {action.type}"
            )
            continue

        path, sheet_name, _ = _target(session, action.sheet)
        current_names = names_for(path)
        invalid_range = False
        shapes: dict[str, tuple[int, int]] = {}
        for label, address in _action_range_fields(action):
            try:
                shapes[address] = _preflight_range_shape(address)
            except (TypeError, ValueError) as error:
                issues.append(
                    f"第 {index + 1} 步：{label}「{address}」无效（{error}）"
                )
                invalid_range = True
        if invalid_range:
            continue

        if action.type in {"writeValues", "writeFormulas"}:
            matrix = (
                action.values
                if action.type == "writeValues"
                else action.formulas
            )
            expected_shape = (len(matrix), len(matrix[0]))
            if shapes[action.range] != expected_shape:
                rows, columns = shapes[action.range]
                issues.append(
                    f"第 {index + 1} 步：目标区域 {action.range} 是"
                    f" {rows}×{columns}，但写入矩阵是"
                    f" {expected_shape[0]}×{expected_shape[1]}"
                )
        if action.type == "writeTable" and shapes[action.startCell] != (1, 1):
            issues.append(
                f"第 {index + 1} 步：表格起始位置必须是单个单元格"
            )
        if action.type == "sortRange":
            columns = shapes[action.range][1]
            for key in action.keys:
                if key.column >= columns:
                    issues.append(
                        f"第 {index + 1} 步：排序列索引 {key.column} "
                        f"超出 {action.range} 的 {columns} 列范围"
                    )
        if action.type == "filterRange":
            columns = shapes[action.range][1]
            if action.column >= columns:
                issues.append(
                    f"第 {index + 1} 步：筛选列索引 {action.column} "
                    f"超出 {action.range} 的 {columns} 列范围"
                )

        if action.type in {"createTable", "addNamedRange"}:
            object_name = (
                action.name
                if action.type == "addNamedRange" or action.name
                else f"ExcelBroTable{index + 1}"
            )
            names = objects_for(path)
            if object_name.casefold() in names:
                issues.append(
                    f"第 {index + 1} 步：表格或命名区域名称"
                    f"「{object_name}」已存在"
                )
            else:
                names.add(object_name.casefold())

        if action.type == "deleteWorksheet":
            if sheet_name not in current_names:
                issues.append(
                    f"第 {index + 1} 步：未找到工作表「{sheet_name}」"
                )
            elif len(current_names) == 1:
                issues.append(
                    f"第 {index + 1} 步：不能删除工作簿中唯一的工作表"
                )
            else:
                current_names.remove(sheet_name)
            continue

        if (
            action.type == "splitGroupAggregate"
            and sheet_name not in current_names
        ):
            issues.append(
                f"第 {index + 1} 步：未找到源工作表「{sheet_name}」"
            )
            continue

        if action.type == "copyRange":
            source_path, source_name, _ = _target(
                session, action.sourceSheet
            )
            if source_name not in names_for(source_path):
                issues.append(
                    f"第 {index + 1} 步：未找到复制源工作表"
                    f"「{action.sourceSheet}」"
                )
                continue

        if action.type != "splitGroupAggregate":
            current_names.add(sheet_name)

    if issues:
        raise ValueError("执行前检查未通过：" + "；".join(issues))


def _verify_folder_plan(
    session: _FolderSession, plan: AnalysisPlan
) -> VerificationReport:
    readers: dict[Path, Workbook] = {}
    checks: list[VerificationCheck] = []
    try:
        for criterion in plan.acceptanceCriteria:
            path, sheet_name, _ = _target(session, criterion.sheet)
            if not path.exists():
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=False,
                        message=f"文件不存在：{path.name}",
                    )
                )
                continue

            workbook = readers.setdefault(
                path,
                load_workbook(
                    path,
                    data_only=False,
                    keep_vba=path.suffix.lower() == ".xlsm",
                ),
            )
            if criterion.type == "worksheetMissing":
                missing = sheet_name not in workbook.sheetnames
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=missing,
                        message=(
                            f"工作表「{sheet_name}」已删除"
                            if missing
                            else f"工作表「{sheet_name}」仍然存在"
                        ),
                    )
                )
                continue
            if sheet_name not in workbook.sheetnames:
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=False,
                        message=f"未找到工作表「{sheet_name}」",
                    )
                )
                continue

            if criterion.type == "worksheetExists":
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=True,
                        message=f"工作表「{sheet_name}」存在",
                    )
                )
                continue

            sheet = workbook[sheet_name]
            if criterion.type == "filterApplied":
                filter_columns = {
                    item.colId: item for item in sheet.auto_filter.filterColumn
                }
                filter_column = filter_columns.get(criterion.column)
                actual_values = (
                    list(filter_column.filters.filter)
                    if filter_column is not None
                    and filter_column.filters is not None
                    else []
                )
                passed = (
                    _normalized_range(sheet.auto_filter.ref or "")
                    == _normalized_range(criterion.range)
                    and sorted(str(value) for value in actual_values)
                    == sorted(str(value) for value in criterion.values)
                )
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"「{sheet_name}」{criterion.range} 第 "
                            f"{criterion.column + 1} 列筛选条件一致"
                            if passed
                            else f"「{sheet_name}」{criterion.range} 的筛选范围或条件与预期不一致"
                        ),
                    )
                )
                continue
            if criterion.type == "filterCleared":
                passed = not sheet.auto_filter.filterColumn
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"工作表「{sheet_name}」的筛选条件已清除"
                            if passed
                            else f"工作表「{sheet_name}」仍有筛选条件"
                        ),
                    )
                )
                continue
            if criterion.type == "tableExists":
                expected_range = _normalized_range(criterion.range)
                matching_table = next(
                    (
                        table
                        for table in sheet.tables.values()
                        if (
                            criterion.name is None
                            or table.name.casefold() == criterion.name.casefold()
                        )
                        and _normalized_range(table.ref) == expected_range
                    ),
                    None,
                )
                actual_headers = (
                    matching_table is not None
                    and getattr(matching_table, "headerRowCount", 1) != 0
                )
                passed = (
                    matching_table is not None
                    and actual_headers == criterion.hasHeaders
                )
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"「{sheet_name}」{criterion.range} 的表格范围与表头状态一致"
                            if passed
                            else f"「{sheet_name}」{criterion.range} 未找到符合预期的表格"
                        ),
                    )
                )
                continue
            if criterion.type == "rangeFormatMatches":
                passed = _range_format_matches(sheet, criterion)
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"「{sheet_name}」{criterion.range} 的格式属性一致"
                            if passed
                            else f"「{sheet_name}」{criterion.range} 的格式属性与预期不一致"
                        ),
                    )
                )
                continue
            if criterion.type == "bordersMatch":
                passed = _borders_match(sheet, criterion)
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"「{sheet_name}」{criterion.range} 的边框属性一致"
                            if passed
                            else f"「{sheet_name}」{criterion.range} 的边框属性与预期不一致"
                        ),
                    )
                )
                continue
            if criterion.type == "dataValidationMatches":
                passed = _data_validation_matches(sheet, criterion)
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"「{sheet_name}」{criterion.range} 的数据验证规则一致"
                            if passed
                            else f"「{sheet_name}」{criterion.range} 的数据验证规则与预期不一致"
                        ),
                    )
                )
                continue
            if criterion.type == "freezePanesMatches":
                actual_rows, actual_columns = _freeze_panes_position(sheet)
                passed = (
                    actual_rows == criterion.rows
                    and actual_columns == criterion.columns
                )
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=passed,
                        message=(
                            f"工作表「{sheet_name}」的冻结窗格位置一致"
                            if passed
                            else f"工作表「{sheet_name}」的冻结窗格位置与预期不一致"
                        ),
                    )
                )
                continue
            if criterion.type in {"chartExists", "pivotTableExists"}:
                checks.append(
                    VerificationCheck(
                        criterion=criterion,
                        passed=False,
                        message=(
                            "文件夹模式当前不能独立验证"
                            + (
                                "图表"
                                if criterion.type == "chartExists"
                                else "数据透视表"
                            )
                            + "对象"
                        ),
                    )
                )
                continue

            min_column, min_row, max_column, max_row = range_boundaries(
                criterion.range
            )
            actual = [
                [
                    sheet.cell(row=row, column=column).value
                    for column in range(min_column, max_column + 1)
                ]
                for row in range(min_row, max_row + 1)
            ]
            if criterion.type == "rangeEmpty":
                passed = all(
                    value in (None, "") for row in actual for value in row
                )
            elif criterion.type == "rangeSorted":
                passed = _range_is_sorted(
                    actual, criterion.keys, criterion.hasHeaders
                )
            else:
                passed = len(actual) == len(criterion.expected) and all(
                    len(actual_row) == len(expected_row)
                    and all(
                        _values_equal(actual_value, expected_value)
                        for actual_value, expected_value in zip(
                            actual_row, expected_row
                        )
                    )
                    for actual_row, expected_row in zip(actual, criterion.expected)
                )
            checks.append(
                VerificationCheck(
                    criterion=criterion,
                    passed=passed,
                    message=(
                        (
                            f"「{sheet_name}」{criterion.range} 已清空"
                            if criterion.type == "rangeEmpty"
                            else (
                                f"「{sheet_name}」{criterion.range} 排序顺序一致"
                                if criterion.type == "rangeSorted"
                                else (
                                    f"「{sheet_name}」{criterion.range} 公式一致"
                                    if criterion.type == "formulasEqual"
                                    else f"「{sheet_name}」{criterion.range} 写入值一致"
                                )
                            )
                        )
                        if passed
                        else f"「{sheet_name}」{criterion.range} 与预期不一致"
                    ),
                    actual=actual,
                )
            )
    finally:
        for workbook in readers.values():
            workbook.close()

    def same_range(criterion, sheet: str, address: str) -> bool:
        return (
            hasattr(criterion, "range")
            and criterion.sheet.casefold() == sheet.casefold()
            and _normalized_range(criterion.range) == _normalized_range(address)
        )

    def has_criterion(predicate) -> bool:
        return any(predicate(criterion) for criterion in plan.acceptanceCriteria)

    unverified_actions: list[UnverifiedAction] = []
    for index, action in enumerate(plan.actions):
        verified = False
        later_actions = plan.actions[index + 1 :]

        def later_same_range(action_type: str):
            if not hasattr(action, "range"):
                return []
            return [
                later
                for later in later_actions
                if later.type == action_type
                and later.sheet.casefold() == action.sheet.casefold()
                and hasattr(later, "range")
                and _normalized_range(later.range)
                == _normalized_range(action.range)
            ]

        if action.type == "createWorksheet":
            verified = has_criterion(
                lambda criterion: criterion.type == "worksheetExists"
                and criterion.sheet.casefold() == action.sheet.casefold()
            )
        elif action.type == "deleteWorksheet":
            verified = has_criterion(
                lambda criterion: criterion.type == "worksheetMissing"
                and criterion.sheet.casefold() == action.sheet.casefold()
            )
        elif action.type == "writeValues":
            verified = has_criterion(
                lambda criterion: criterion.type == "rangeEquals"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "writeFormulas":
            verified = has_criterion(
                lambda criterion: criterion.type == "formulasEqual"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "writeTable":
            start_column, start_row, _, _ = range_boundaries(
                action.startCell
            )
            target_range = (
                f"{get_column_letter(start_column)}{start_row}:"
                f"{get_column_letter(start_column + len(action.headers) - 1)}"
                f"{start_row + len(action.rows)}"
            )
            verified = has_criterion(
                lambda criterion: criterion.type == "rangeEquals"
                and same_range(criterion, action.sheet, target_range)
            )
        elif action.type == "clearRange":
            verified = action.applyTo in {"all", "contents"} and has_criterion(
                lambda criterion: criterion.type == "rangeEmpty"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "copyRange":
            verified = has_criterion(
                lambda criterion: criterion.type
                in {"rangeEquals", "formulasEqual"}
                and same_range(criterion, action.sheet, action.targetRange)
            )
        elif action.type == "sortRange":
            verified = has_criterion(
                lambda criterion: criterion.type == "rangeSorted"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "filterRange":
            verified = any(
                later.type in {"filterRange", "clearFilter"}
                and later.sheet.casefold() == action.sheet.casefold()
                for later in later_actions
            ) or has_criterion(
                lambda criterion: criterion.type == "filterApplied"
                and same_range(criterion, action.sheet, action.range)
                and criterion.column == action.column
            )
        elif action.type == "clearFilter":
            verified = any(
                later.type in {"filterRange", "clearFilter"}
                and later.sheet.casefold() == action.sheet.casefold()
                for later in later_actions
            ) or has_criterion(
                lambda criterion: criterion.type == "filterCleared"
                and criterion.sheet.casefold() == action.sheet.casefold()
            )
        elif action.type == "createTable":
            verified = has_criterion(
                lambda criterion: criterion.type == "tableExists"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "setFill":
            verified = bool(later_same_range("setFill")) or has_criterion(
                lambda criterion: criterion.type == "rangeFormatMatches"
                and same_range(criterion, action.sheet, action.range)
                and criterion.fillColor is not None
                and criterion.fillColor.casefold() == action.color.casefold()
            )
        elif action.type == "setFont":
            verified = (
                action.bold is None
                or any(
                    later.bold is not None
                    for later in later_same_range("setFont")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.bold == action.bold
                )
            ) and (
                action.color is None
                or any(
                    later.color is not None
                    for later in later_same_range("setFont")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.fontColor is not None
                    and criterion.fontColor.casefold() == action.color.casefold()
                )
            )
        elif action.type == "setNumberFormat":
            verified = bool(
                later_same_range("setNumberFormat")
            ) or has_criterion(
                lambda criterion: criterion.type == "rangeFormatMatches"
                and same_range(criterion, action.sheet, action.range)
                and criterion.numberFormat == action.formatCode
            )
        elif action.type == "setAlignment":
            verified = (
                action.horizontal is None
                or any(
                    later.horizontal is not None
                    for later in later_same_range("setAlignment")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.horizontal == action.horizontal
                )
            ) and (
                action.vertical is None
                or any(
                    later.vertical is not None
                    for later in later_same_range("setAlignment")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.vertical == action.vertical
                )
            ) and (
                action.wrapText is None
                or any(
                    later.wrapText is not None
                    for later in later_same_range("setAlignment")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.wrapText == action.wrapText
                )
            )
        elif action.type == "resizeRange":
            verified = (
                action.rowHeight is None
                or any(
                    later.rowHeight is not None
                    for later in later_same_range("resizeRange")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.rowHeight == action.rowHeight
                )
            ) and (
                action.columnWidth is None
                or any(
                    later.columnWidth is not None
                    for later in later_same_range("resizeRange")
                )
                or has_criterion(
                    lambda criterion: criterion.type == "rangeFormatMatches"
                    and same_range(criterion, action.sheet, action.range)
                    and criterion.columnWidth == action.columnWidth
                )
            )
        elif action.type == "setBorders":
            verified = has_criterion(
                lambda criterion: criterion.type == "bordersMatch"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "setDataValidation":
            verified = bool(
                later_same_range("setDataValidation")
            ) or has_criterion(
                lambda criterion: criterion.type == "dataValidationMatches"
                and same_range(criterion, action.sheet, action.range)
            )
        elif action.type == "freezePanes":
            verified = any(
                later.type == "freezePanes"
                and later.sheet.casefold() == action.sheet.casefold()
                for later in later_actions
            ) or has_criterion(
                lambda criterion: criterion.type == "freezePanesMatches"
                and criterion.sheet.casefold() == action.sheet.casefold()
                and criterion.rows == action.rows
                and criterion.columns == action.columns
            )
        elif action.type == "createChart":
            verified = has_criterion(
                lambda criterion: criterion.type == "chartExists"
                and criterion.sheet.casefold() == action.sheet.casefold()
                and _normalized_range(criterion.sourceRange)
                == _normalized_range(action.sourceRange)
            )
        elif action.type == "createPivotTable":
            verified = has_criterion(
                lambda criterion: criterion.type == "pivotTableExists"
                and criterion.sheet.casefold() == action.sheet.casefold()
                and criterion.name.casefold() == action.name.casefold()
            )
        if not verified:
            unverified_actions.append(
                UnverifiedAction(
                    index=index,
                    type=action.type,
                    sheet=action.sheet,
                    message=(
                        f"第 {index + 1} 步 {action.type} 已执行，"
                        "但当前验收协议不能独立验证该操作的具体效果"
                    ),
                )
            )

    if any(not check.passed for check in checks):
        status = "failed"
    elif unverified_actions or not checks:
        status = "executed_unverified"
    else:
        status = "verified"
    return VerificationReport(
        status=status,
        passed=status == "verified",
        checks=checks,
        unverifiedActions=unverified_actions,
    )


def _save_workbooks_atomically(
    session: _FolderSession,
    workbooks: dict[Path, Workbook],
    modified_paths: set[Path],
) -> tuple[list[str], list[str]]:
    targets = [
        (path, workbooks[path])
        for path in workbooks
        if path in modified_paths
    ]
    temporary_paths: dict[Path, Path] = {}
    existed_before = {path: path.exists() for path, _ in targets}
    try:
        for path, workbook in targets:
            temporary = path.with_name(
                f".{path.name}.excel-bro-tmp-{uuid.uuid4().hex}{path.suffix}"
            )
            workbook.save(temporary)
            temporary_paths[path] = temporary
    except Exception:
        for temporary in temporary_paths.values():
            temporary.unlink(missing_ok=True)
        raise
    finally:
        for workbook in workbooks.values():
            workbook.close()

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backups_by_path: dict[Path, Path] = {}
    try:
        for path, _ in targets:
            if not existed_before[path]:
                continue
            backup = path.with_name(
                f"{path.name}.excel-bro-backup-{timestamp}-"
                f"{uuid.uuid4().hex[:8]}{path.suffix}"
            )
            shutil.copy2(path, backup)
            backups_by_path[path] = backup
    except Exception:
        for temporary in temporary_paths.values():
            temporary.unlink(missing_ok=True)
        raise

    replaced: list[Path] = []
    try:
        for path, _ in targets:
            temporary_paths[path].replace(path)
            replaced.append(path)
    except Exception as replace_error:
        rollback_errors: list[str] = []
        for path in reversed(replaced):
            try:
                backup = backups_by_path.get(path)
                if backup is not None:
                    shutil.copy2(backup, path)
                elif not existed_before[path]:
                    path.unlink(missing_ok=True)
            except Exception as rollback_error:
                rollback_errors.append(f"{path.name}: {rollback_error}")
        for path, temporary in temporary_paths.items():
            if path not in replaced:
                temporary.unlink(missing_ok=True)
        detail = (
            f"；恢复失败：{'；'.join(rollback_errors)}"
            if rollback_errors
            else ""
        )
        raise RuntimeError(
            f"替换工作簿失败，已恢复先前文件：{replace_error}{detail}"
        ) from replace_error

    modified = [str(path.relative_to(session.root)) for path, _ in targets]
    backups = [
        str(backups_by_path[path].relative_to(session.root))
        for path, _ in targets
        if path in backups_by_path
    ]
    return modified, backups


def _execute_folder_plan(request: FolderExecuteRequest) -> FolderExecuteResponse:
    session = _session(request.sessionId)
    if session.source_fingerprint is not None:
        if request.plan.sourceFingerprint is None:
            raise ValueError("计划缺少数据来源指纹，请重新生成预览")
        if request.plan.sourceFingerprint != session.source_fingerprint:
            raise ValueError("计划的数据来源指纹已失效，请重新生成预览")
        current_fingerprint = _selected_files_fingerprint(
            {path for path, _ in session.selected_targets.values()}
        )
        if current_fingerprint != session.source_fingerprint:
            raise ValueError("预览后所选文件已发生变化，请重新生成预览后再执行")
    _preflight_folder_plan(session, request.plan)
    execution_started = time.perf_counter()
    workbooks: dict[Path, Workbook] = {}
    modified_paths: set[Path] = set()
    action_results: list[ActionExecutionResult] = []

    for index, action in enumerate(request.plan.actions):
        path, sheet_name, _ = _target(session, action.sheet)
        workbook = workbooks.setdefault(path, _load_for_write(path))
        modified_paths.add(path)

        if action.type == "deleteWorksheet":
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"未找到工作表「{sheet_name}」")
            if len(workbook.sheetnames) == 1:
                raise ValueError("不能删除工作簿中唯一的工作表")
            workbook.remove(workbook[sheet_name])
            action_results.append(
                ActionExecutionResult(
                    index=index,
                    type=action.type,
                    sheet=action.sheet,
                    status="succeeded",
                )
            )
            continue

        if action.type == "splitGroupAggregate":
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"未找到源工作表「{sheet_name}」")
            source_sheet = workbook[sheet_name]
            if action.sourceRange:
                min_col, min_row, max_col, max_row = range_boundaries(
                    action.sourceRange
                )
            else:
                min_col, min_row = 1, 1
                max_col, max_row = source_sheet.max_column, source_sheet.max_row
            values = [
                [
                    source_sheet.cell(row=row, column=column).value
                    for column in range(min_col, max_col + 1)
                ]
                for row in range(min_row, max_row + 1)
            ]

            def normalized_header(value) -> str:
                return str(value or "").strip().casefold()

            required_fields = [
                action.splitBy,
                *action.groupBy,
                *[
                    metric.field
                    for metric in action.metrics
                    if metric.field is not None
                ],
            ]
            required_normalized = {
                normalized_header(field) for field in required_fields
            }
            header_offset = next(
                (
                    row_index
                    for row_index, row in enumerate(values)
                    if required_normalized.issubset(
                        {
                            normalized_header(value)
                            for value in row
                            if normalized_header(value)
                        }
                    )
                ),
                None,
            )
            if header_offset is None:
                raise ValueError(
                    f"未找到包含这些字段的表头：{'、'.join(required_fields)}"
                )
            header_indexes: dict[str, int] = {}
            for column_index, value in enumerate(values[header_offset]):
                normalized = normalized_header(value)
                if normalized and normalized not in header_indexes:
                    header_indexes[normalized] = column_index

            def field_index(field: str) -> int:
                try:
                    return header_indexes[normalized_header(field)]
                except KeyError as error:
                    raise ValueError(f"未找到字段「{field}」") from error

            split_index = field_index(action.splitBy)
            group_indexes = [field_index(field) for field in action.groupBy]
            metric_indexes = [
                field_index(metric.field) if metric.field else None
                for metric in action.metrics
            ]

            def is_blank(value) -> bool:
                return value is None or (
                    isinstance(value, str) and not value.strip()
                )

            def value_key(value):
                return (type(value).__name__, str(value).strip())

            groups: dict[tuple, dict] = {}
            splits: dict[tuple, dict] = {}
            for row in values[header_offset + 1 :]:
                split_value = row[split_index]
                group_values = [row[column] for column in group_indexes]
                if (
                    not action.includeBlankSplitValues
                    and is_blank(split_value)
                ) or any(is_blank(value) for value in group_values):
                    continue
                group_key = tuple(value_key(value) for value in group_values)
                split_key = (value_key(split_value),)
                group = groups.setdefault(
                    group_key,
                    {
                        "values": group_values,
                        "totals": [0.0 for _ in action.metrics],
                    },
                )
                split = splits.setdefault(
                    split_key, {"value": split_value, "pairs": {}}
                )
                pair = split["pairs"].setdefault(
                    group_key, [0.0 for _ in action.metrics]
                )
                for metric_index, metric in enumerate(action.metrics):
                    cell = (
                        None
                        if metric_indexes[metric_index] is None
                        else row[metric_indexes[metric_index]]
                    )
                    if metric.operation == "countRows":
                        increment = 1.0
                    elif metric.operation == "countNonBlank":
                        increment = 0.0 if is_blank(cell) else 1.0
                    else:
                        try:
                            increment = float(
                                str(cell or "").replace(",", "").strip()
                            )
                        except ValueError:
                            increment = 0.0
                    group["totals"][metric_index] += increment
                    pair[metric_index] += increment

            if not splits:
                raise ValueError("表头下没有可拆分的有效数据行")
            if len(splits) > action.maxOutputSheets:
                raise ValueError(
                    f"将生成 {len(splits)} 个工作表，"
                    f"超过安全上限 {action.maxOutputSheets} 个"
                )

            headers = [*action.groupBy, action.splitBy]
            ratio_columns: list[int] = []
            for metric in action.metrics:
                headers.append(metric.outputName)
                if metric.ratioOutputName:
                    headers.append(metric.ratioOutputName)
                    ratio_columns.append(len(headers))

            existing_names = {
                name.casefold(): name for name in workbook.sheetnames
            }
            generated_names: set[str] = set()

            def safe_name(value) -> str:
                cleaned = re.sub(
                    r"\s+",
                    " ",
                    re.sub(
                        r"[\x00-\x1f:：\\＼/／?？*＊\[\]［］]",
                        " ",
                        str(value or ""),
                    ),
                ).strip().strip("'").strip()
                cleaned = cleaned or "未命名"
                if cleaned.casefold() == "history":
                    cleaned = "History 结果"
                return cleaned[:31]

            def renamed(base: str) -> str:
                if base.casefold() not in existing_names:
                    return base
                suffix = 2
                while True:
                    marker = f" ({suffix})"
                    candidate = f"{base[:31 - len(marker)]}{marker}"
                    if candidate.casefold() not in existing_names:
                        return candidate
                    suffix += 1

            for split in splits.values():
                base_name = safe_name(split["value"])
                existing_name = existing_names.get(base_name.casefold())
                collides_with_current_run = base_name.casefold() in generated_names
                if (
                    existing_name
                    and action.existingSheetPolicy == "skip"
                    and not collides_with_current_run
                ):
                    continue
                if (
                    existing_name
                    and action.existingSheetPolicy == "replace"
                    and not collides_with_current_run
                ):
                    if existing_name.casefold() == sheet_name.casefold():
                        raise ValueError(
                            f"不能用拆分结果覆盖源工作表「{sheet_name}」"
                        )
                    position = workbook.sheetnames.index(existing_name)
                    workbook.remove(workbook[existing_name])
                    target_sheet = workbook.create_sheet(base_name, position)
                    target_name = base_name
                else:
                    target_name = renamed(base_name)
                    target_sheet = workbook.create_sheet(target_name)
                existing_names[target_name.casefold()] = target_name
                generated_names.add(target_name.casefold())

                target_sheet.append(headers)
                for group_key, aggregates in split["pairs"].items():
                    group = groups[group_key]
                    output_row = [*group["values"], split["value"]]
                    for metric_index, metric in enumerate(action.metrics):
                        aggregate = aggregates[metric_index]
                        if metric.operation != "sum" and aggregate.is_integer():
                            aggregate = int(aggregate)
                        output_row.append(aggregate)
                        if metric.ratioOutputName:
                            total = group["totals"][metric_index]
                            output_row.append(
                                None if total == 0 else aggregate / total
                            )
                    target_sheet.append(output_row)
                for cell in target_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DFF3E4")
                for column in ratio_columns:
                    for row in range(2, target_sheet.max_row + 1):
                        target_sheet.cell(row=row, column=column).number_format = (
                            "0.00%"
                        )
                for column in range(1, target_sheet.max_column + 1):
                    letter = get_column_letter(column)
                    width = max(
                        len(str(target_sheet.cell(row=row, column=column).value or ""))
                        for row in range(1, target_sheet.max_row + 1)
                    )
                    target_sheet.column_dimensions[letter].width = min(width + 2, 40)

            action_results.append(
                ActionExecutionResult(
                    index=index,
                    type=action.type,
                    sheet=action.sheet,
                    status="succeeded",
                )
            )
            continue

        sheet = _get_or_create_sheet(workbook, sheet_name)
        if action.type in {"createWorksheet", "activateWorksheet"}:
            if action.type == "activateWorksheet":
                workbook.active = workbook.sheetnames.index(sheet.title)
            action_results.append(
                ActionExecutionResult(
                    index=index,
                    type=action.type,
                    sheet=action.sheet,
                    status="succeeded",
                )
            )
            continue
        if action.type == "writeTable":
            rows = [action.headers, *action.rows]
            start_column, start_row, _, _ = range_boundaries(action.startCell)
            for row_offset, row in enumerate(rows):
                for column_offset, value in enumerate(row):
                    sheet.cell(
                        row=start_row + row_offset,
                        column=start_column + column_offset,
                        value=value,
                    )
        elif action.type == "writeValues":
            _write_matrix(sheet, action.range, action.values)
        elif action.type == "writeFormulas":
            _write_matrix(sheet, action.range, action.formulas)
        elif action.type == "clearRange":
            if action.applyTo in {"all", "contents"}:
                for row in _iter_range(sheet, action.range):
                    for cell in row:
                        cell.value = None
            if action.applyTo in {"all", "formats"}:
                for row in _iter_range(sheet, action.range):
                    for cell in row:
                        cell._style = copy(sheet["A1"]._style)
                        cell.number_format = "General"
                        cell.alignment = Alignment()
            if action.applyTo in {"all", "hyperlinks"}:
                for row in _iter_range(sheet, action.range):
                    for cell in row:
                        cell.hyperlink = None
        elif action.type in {"insertRange", "deleteRange"}:
            _folder_mode_unsupported(action.type)
        elif action.type == "copyRange":
            source_path, source_name, _ = _target(session, action.sourceSheet)
            source_book = workbooks.setdefault(source_path, _load_for_write(source_path))
            if source_name not in source_book.sheetnames:
                raise ValueError(f"未找到源工作表「{source_name}」")
            source_sheet = source_book[source_name]
            source_cells = [list(row) for row in _iter_range(source_sheet, action.sourceRange)]
            target_min_col, target_min_row, _, _ = range_boundaries(action.targetRange)
            if action.transpose:
                source_cells = [list(row) for row in zip(*source_cells)]
            for row_offset, row in enumerate(source_cells):
                for column_offset, source_cell in enumerate(row):
                    target_cell = sheet.cell(
                        row=target_min_row + row_offset,
                        column=target_min_col + column_offset,
                    )
                    if action.skipBlanks and source_cell.value in (None, ""):
                        continue
                    if action.copyType in {"all", "values", "formulas"}:
                        target_cell.value = source_cell.value
                    if action.copyType in {"all", "formats"}:
                        target_cell._style = copy(source_cell._style)
                        target_cell.number_format = source_cell.number_format
                    if action.copyType == "link":
                        target_cell.value = (
                            f"='{source_name}'!{source_cell.coordinate}"
                        )
        elif action.type == "sortRange":
            min_col, min_row, max_col, max_row = range_boundaries(action.range)
            first_data_row = min_row + (1 if action.hasHeaders else 0)
            rows = [
                [sheet.cell(row=row, column=col).value for col in range(min_col, max_col + 1)]
                for row in range(first_data_row, max_row + 1)
            ]
            if rows and any(key.column >= len(rows[0]) for key in action.keys):
                raise ValueError("排序列索引超出范围")
            rows.sort(
                key=cmp_to_key(
                    lambda left, right: _compare_sorted_rows(
                        left, right, action.keys
                    )
                )
            )
            for row_offset, values in enumerate(rows):
                for column_offset, value in enumerate(values):
                    sheet.cell(
                        row=first_data_row + row_offset,
                        column=min_col + column_offset,
                        value=value,
                    )
        elif action.type == "filterRange":
            sheet.auto_filter.ref = action.range
            sheet.auto_filter.filterColumn = [
                FilterColumn(
                    colId=action.column,
                    filters=Filters(filter=[str(value) for value in action.values]),
                )
            ]
        elif action.type == "clearFilter":
            sheet.auto_filter.ref = None
            sheet.auto_filter.filterColumn = []
        elif action.type == "setDataValidation":
            operator_map = {
                "equalTo": "equal",
                "notEqualTo": "notEqual",
                "greaterThanOrEqualTo": "greaterThanOrEqual",
                "lessThanOrEqualTo": "lessThanOrEqual",
            }
            validation_type = {
                "wholeNumber": "whole",
                "textLength": "textLength",
            }.get(action.validationType, action.validationType)
            if validation_type == "list":
                formula1 = '"' + ",".join(str(value) for value in action.values) + '"'
            else:
                formula1 = action.formula1
            validation = DataValidation(
                type=validation_type,
                operator=operator_map.get(action.operator, action.operator),
                formula1=formula1,
                formula2=action.formula2,
                allow_blank=action.allowBlank,
                prompt=action.prompt,
                error=action.errorMessage,
            )
            sheet.add_data_validation(validation)
            validation.add(action.range)
        elif action.type == "setConditionalFormat":
            fill = PatternFill(
                "solid", fgColor=(action.color or "#FFF2CC").removeprefix("#")
            )
            if action.ruleType == "colorScale":
                rule = ColorScaleRule(
                    start_type="min",
                    start_color=(action.minColor or "#F8696B").removeprefix("#"),
                    mid_type="percentile" if action.midColor else None,
                    mid_value=50 if action.midColor else None,
                    mid_color=action.midColor.removeprefix("#") if action.midColor else None,
                    end_type="max",
                    end_color=(action.maxColor or "#63BE7B").removeprefix("#"),
                )
            elif action.ruleType == "custom":
                rule = FormulaRule(formula=[str(action.formula1 or "")], fill=fill)
            else:
                rule = CellIsRule(
                    operator={
                        "equalTo": "equal",
                        "notEqualTo": "notEqual",
                        "greaterThanOrEqualTo": "greaterThanOrEqual",
                        "lessThanOrEqualTo": "lessThanOrEqual",
                    }.get(action.operator or "equalTo", action.operator or "equal"),
                    formula=[
                        str(action.formula1 or ""),
                        *(
                            [str(action.formula2)]
                            if action.formula2 is not None
                            else []
                        ),
                    ],
                    fill=fill,
                )
            sheet.conditional_formatting.add(action.range, rule)
        elif action.type == "setNumberFormat":
            for row in _iter_range(sheet, action.range):
                for cell in row:
                    cell.number_format = action.formatCode
        elif action.type == "setBorders":
            side = Side(
                style=_openpyxl_border_style(action.style, action.weight),
                color=action.color.removeprefix("#"),
            )
            for row in _iter_range(sheet, action.range):
                for cell in row:
                    current = cell.border
                    cell.border = Border(
                        left=side if "left" in action.sides else current.left,
                        right=side if "right" in action.sides else current.right,
                        top=side if "top" in action.sides else current.top,
                        bottom=side if "bottom" in action.sides else current.bottom,
                        vertical=side if "insideVertical" in action.sides else current.vertical,
                        horizontal=side if "insideHorizontal" in action.sides else current.horizontal,
                    )
        elif action.type == "setAlignment":
            for row in _iter_range(sheet, action.range):
                for cell in row:
                    cell.alignment = copy(cell.alignment)
                    if action.horizontal is not None:
                        cell.alignment = Alignment(
                            horizontal={
                                "centerAcrossSelection": "centerContinuous"
                            }.get(action.horizontal, action.horizontal),
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                        )
                    if action.vertical is not None or action.wrapText is not None:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=action.vertical or cell.alignment.vertical,
                            wrap_text=(
                                action.wrapText
                                if action.wrapText is not None
                                else cell.alignment.wrap_text
                            ),
                        )
        elif action.type == "mergeCells":
            sheet.merge_cells(action.range)
        elif action.type == "unmergeCells":
            sheet.unmerge_cells(action.range)
        elif action.type == "resizeRange":
            min_col, min_row, max_col, max_row = range_boundaries(action.range)
            if action.rowHeight is not None:
                for row in range(min_row, max_row + 1):
                    sheet.row_dimensions[row].height = action.rowHeight
            if action.columnWidth is not None:
                for column in range(min_col, max_col + 1):
                    sheet.column_dimensions[get_column_letter(column)].width = action.columnWidth
        elif action.type == "freezePanes":
            if action.rows == 0 and action.columns == 0:
                sheet.freeze_panes = None
            else:
                sheet.freeze_panes = sheet.cell(
                    row=action.rows + 1, column=action.columns + 1
                )
        elif action.type == "setHyperlink":
            for row in _iter_range(sheet, action.range):
                for cell in row:
                    cell.hyperlink = action.address
                    if action.text is not None:
                        cell.value = action.text
        elif action.type in {"addComment", "addNote"}:
            sheet[action.cell].comment = Comment(action.text, "Excel Bro")
        elif action.type == "createTable":
            table = Table(
                displayName=action.name or f"ExcelBroTable{index + 1}",
                ref=action.range,
            )
            if not action.hasHeaders:
                table.headerRowCount = 0
            table.tableStyleInfo = TableStyleInfo(
                name=action.style or "TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        elif action.type == "createChart":
            chart_class = {
                "ColumnClustered": BarChart,
                "BarClustered": BarChart,
                "Line": LineChart,
                "Pie": PieChart,
                "Doughnut": DoughnutChart,
                "XYScatter": ScatterChart,
                "Area": AreaChart,
            }.get(action.chartType)
            if chart_class is None:
                _folder_mode_unsupported(action.type)
            chart = chart_class()
            if action.chartType == "ColumnClustered":
                chart.type = "col"
            min_col, min_row, max_col, max_row = range_boundaries(action.sourceRange)
            chart.add_data(
                Reference(
                    sheet,
                    min_col=min_col,
                    max_col=max_col,
                    min_row=min_row,
                    max_row=max_row,
                ),
                titles_from_data=True,
            )
            if action.title:
                chart.title = action.title
            sheet.add_chart(chart, action.targetRange or f"{get_column_letter(max_col + 2)}{min_row}")
        elif action.type == "createPivotTable":
            _folder_mode_unsupported(action.type)
        elif action.type == "addNamedRange":
            workbook.defined_names.add(
                DefinedName(
                    action.name,
                    attr_text=(
                        f"{quote_sheetname(sheet.title)}!"
                        f"{absolute_coordinate(action.range)}"
                    ),
                    comment=action.comment,
                )
            )
        elif action.type == "addImage":
            try:
                from openpyxl.drawing.image import Image
                image = Image(io.BytesIO(base64.b64decode(action.base64)))
            except Exception as error:
                raise ValueError(f"无法解析图片数据：{error}") from error
            sheet.add_image(image, action.targetRange.split(":")[0])
        elif action.type == "addShape":
            _folder_mode_unsupported(action.type)
        elif action.type == "setFill":
            min_column, min_row, max_column, max_row = range_boundaries(action.range)
            fill = PatternFill("solid", fgColor=action.color.removeprefix("#"))
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_column,
                max_col=max_column,
            ):
                for cell in row:
                    cell.fill = fill
        elif action.type == "setFont":
            min_column, min_row, max_column, max_row = range_boundaries(action.range)
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_column,
                max_col=max_column,
            ):
                for cell in row:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=action.bold
                        if action.bold is not None
                        else cell.font.bold,
                        italic=cell.font.italic,
                        color=action.color.removeprefix("#")
                        if action.color
                        else cell.font.color,
                    )
        elif action.type == "autofit":
            min_column, _, max_column, _ = range_boundaries(action.range)
            for column in range(min_column, max_column + 1):
                letter = get_column_letter(column)
                width = max(
                    (
                        len(str(cell.value))
                        for cell in sheet[letter]
                        if cell.value is not None
                    ),
                    default=8,
                )
                sheet.column_dimensions[letter].width = min(width + 2, 60)
        action_results.append(
            ActionExecutionResult(
                index=index,
                type=action.type,
                sheet=action.sheet,
                status="succeeded",
            )
        )

    modified, backups = _save_workbooks_atomically(
        session, workbooks, modified_paths
    )

    verification_started = time.perf_counter()
    execution_ms = (verification_started - execution_started) * 1000
    verification = _verify_folder_plan(session, request.plan)
    verification_ms = (time.perf_counter() - verification_started) * 1000
    return FolderExecuteResponse(
        filesModified=modified,
        backups=backups,
        actionResults=action_results,
        verification=verification,
        executionMs=execution_ms,
        verificationMs=verification_ms,
    )


def execute_folder_plan(request: FolderExecuteRequest) -> FolderExecuteResponse:
    opened: list[Workbook] = []
    token = _OPEN_WRITE_WORKBOOKS.set(opened)
    try:
        return _execute_folder_plan(request)
    finally:
        for workbook in opened:
            try:
                workbook.close()
            except Exception:
                pass
        _OPEN_WRITE_WORKBOOKS.reset(token)
