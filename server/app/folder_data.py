from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from .capabilities import capability_int
from .models import CellValue, DataFilter, DataToolRequest, DataToolResult


@dataclass(frozen=True)
class AuthorizedSheet:
    file_id: str
    sheet_id: str
    path: Path
    sheet_name: str
    display_name: str


def _field_key(value: object) -> str:
    return re.sub(r"[\s_\-:：()（）]+", "", str(value or "")).casefold()


def _cell(value: object) -> CellValue:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _number(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    try:
        parsed = float(text)
    except ValueError:
        return None
    return parsed / 100 if percent else parsed


def _compare(actual: object, expected: object) -> int:
    left = _number(actual)
    right = _number(expected)
    if left is not None and right is not None:
        return (left > right) - (left < right)
    left_text = str(_cell(actual) or "").casefold()
    right_text = str(_cell(expected) or "").casefold()
    return (left_text > right_text) - (left_text < right_text)


def _matches(actual: object, filter_: DataFilter) -> bool:
    expected = filter_.value
    if filter_.operator == "isBlank":
        return actual is None or pd.isna(actual) or actual == ""
    if filter_.operator == "isNotBlank":
        return not _matches(actual, filter_.model_copy(update={"operator": "isBlank"}))
    if filter_.operator == "contains":
        return str(expected or "").casefold() in str(_cell(actual) or "").casefold()
    compared = _compare(actual, expected)
    return {
        "equals": compared == 0,
        "notEquals": compared != 0,
        "greaterThan": compared > 0,
        "greaterThanOrEqual": compared >= 0,
        "lessThan": compared < 0,
        "lessThanOrEqual": compared <= 0,
    }[filter_.operator]


def _read_authorized(source: AuthorizedSheet) -> pd.DataFrame:
    workbook = load_workbook(
        source.path,
        read_only=True,
        data_only=True,
        keep_vba=source.path.suffix.lower() == ".xlsm",
    )
    try:
        sheet = workbook[source.sheet_name]
        max_rows = capability_int("queryTable", "maxRows")
        max_columns = capability_int("queryTable", "maxColumns")
        max_cells = capability_int("queryTable", "maxCells")
        if sheet.max_row > max_rows:
            raise ValueError(
                f"「{source.display_name}」有 {sheet.max_row} 行，"
                f"超过 {max_rows} 行安全上限"
            )
        if sheet.max_column > max_columns:
            raise ValueError(
                f"「{source.display_name}」有 {sheet.max_column} 列，"
                f"超过 {max_columns} 列安全上限"
            )
        if sheet.max_row * sheet.max_column > max_cells:
            raise ValueError(
                f"「{source.display_name}」包含约 "
                f"{sheet.max_row * sheet.max_column} 个单元格，"
                f"超过 {max_cells} 个单元格安全上限"
            )

        def normalized(cell):
            value = cell.value
            if isinstance(value, datetime):
                return value.date().isoformat()
            if isinstance(value, date):
                return value.isoformat()
            if (
                isinstance(value, (int, float))
                and re.fullmatch(r"0+", cell.number_format or "")
            ):
                return f"{int(value):0{len(cell.number_format)}d}"
            if isinstance(value, (str, int, float, bool)) or value is None:
                return value
            return str(value)

        rows = [
            [normalized(cell) for cell in row]
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                max_col=sheet.max_column,
            )
        ]
    finally:
        workbook.close()
    if not rows:
        return pd.DataFrame()
    headers = [
        str(value).strip() if value not in (None, "") else f"未命名列{index + 1}"
        for index, value in enumerate(rows[0])
    ]
    frame = pd.DataFrame(rows[1:], columns=headers, dtype=object)
    frame.insert(0, "工作表", source.display_name)
    frame.insert(1, "来源文件ID", source.file_id)
    frame.insert(2, "来源工作表ID", source.sheet_id)
    return frame


def _resolve_columns(
    frame: pd.DataFrame, fields: list[str]
) -> tuple[dict[str, str], list[str]]:
    by_key = {_field_key(column): str(column) for column in frame.columns}
    resolved: dict[str, str] = {}
    missing: list[str] = []
    for field in fields:
        column = by_key.get(_field_key(field))
        if column is None:
            missing.append(field)
        else:
            resolved[field] = column
    return resolved, missing


def execute_folder_query(
    request: DataToolRequest,
    authorized_sheets: list[AuthorizedSheet],
) -> DataToolResult:
    if not authorized_sheets:
        raise ValueError("没有已授权的文件夹工作表")
    arguments = request.arguments
    sources = (
        authorized_sheets[:1]
        if arguments.scope == "active"
        else authorized_sheets
    )
    required = [
        *arguments.fields,
        *arguments.groupBy,
        *[item.field for item in arguments.filters],
        *[item.field for item in arguments.metrics if item.field],
        *([arguments.profileField] if arguments.profileField else []),
        *(arguments.combine.deduplicateBy if arguments.combine else []),
        *(
            [
                arguments.combine.leftKey,
                arguments.combine.rightKey,
            ]
            if arguments.combine and arguments.combine.mode == "join"
            else []
        ),
    ]
    required = [field for field in required if field]
    frames: list[pd.DataFrame] = []
    warnings: list[str] = []
    for source in sources:
        frame = _read_authorized(source)
        if frame.empty:
            warnings.append(f"「{source.display_name}」没有数据行")
            continue
        source_required = required
        if arguments.combine and arguments.combine.mode == "join":
            source_required = (
                [arguments.combine.leftKey]
                if source.sheet_id == arguments.combine.leftSourceSheetId
                else [arguments.combine.rightKey]
                if source.sheet_id == arguments.combine.rightSourceSheetId
                else []
            )
        resolved, missing = _resolve_columns(frame, source_required)
        if missing:
            warnings.append(
                f"「{source.display_name}」未找到字段：{'、'.join(sorted(set(missing)))}"
            )
            continue
        frame = frame.rename(columns={column: field for field, column in resolved.items()})
        frames.append(frame)
    if not frames:
        raise ValueError("；".join(warnings) or "所选工作表没有可查询的数据")

    combine = arguments.combine
    if combine and combine.mode == "join":
        left = next(
            (
                frame
                for frame in frames
                if str(frame["来源工作表ID"].iloc[0])
                == combine.leftSourceSheetId
            ),
            None,
        )
        right = next(
            (
                frame
                for frame in frames
                if str(frame["来源工作表ID"].iloc[0])
                == combine.rightSourceSheetId
            ),
            None,
        )
        if left is None or right is None:
            raise ValueError("关联请求引用了未选择的工作表 ID")
        data = left.merge(
            right,
            left_on=combine.leftKey,
            right_on=combine.rightKey,
            how=combine.joinHow,
            suffixes=("_左", "_右"),
        )
        resolved, missing = _resolve_columns(data, required)
        if missing:
            raise ValueError(f"关联结果未找到字段：{'、'.join(sorted(set(missing)))}")
        data = data.rename(columns={column: field for field, column in resolved.items()})
    else:
        data = pd.concat(frames, ignore_index=True, sort=False)
        if combine and combine.mode == "deduplicate":
            data = data.drop_duplicates(
                subset=combine.deduplicateBy,
                keep="first",
                ignore_index=True,
            )
    max_cells = capability_int("queryTable", "maxCells")
    if data.shape[0] * data.shape[1] > max_cells:
        raise ValueError(
            f"文件夹查询合并后包含约 {data.shape[0] * data.shape[1]} 个单元格，"
            f"超过 {max_cells} 个单元格安全上限"
        )
    max_rows = capability_int("queryTable", "maxRows")
    if len(data) > max_rows:
        raise ValueError(f"文件夹查询包含 {len(data)} 行，超过 {max_rows} 行安全上限")
    scanned_rows = len(data)
    for filter_ in arguments.filters:
        data = data[
            data[filter_.field].map(lambda value, item=filter_: _matches(value, item))
        ]

    if arguments.mode == "profile":
        field = arguments.profileField or ""
        counts = data.groupby(field, dropna=False).size().reset_index(name="数量")
        counts["占比"] = counts["数量"] / max(len(data), 1)
        output = counts
        title = f"{field}分布"
        calculation = f"按「{field}」统计数量，占比 = 该值数量 ÷ 筛选后总记录数。"
    elif arguments.mode == "aggregate":
        group_fields = arguments.groupBy
        if group_fields:
            grouped = data.groupby(group_fields, dropna=False, sort=False)
            output = grouped.size().reset_index(name="__rows")
        else:
            output = pd.DataFrame({"__rows": [len(data)]})
        metric_columns: list[str] = []
        for metric in arguments.metrics:
            if metric.operation == "countRows":
                values = (
                    grouped.size().to_numpy()
                    if group_fields
                    else [len(data)]
                )
            elif metric.operation == "countDistinct":
                values = (
                    grouped[metric.field].nunique(dropna=True).to_numpy()
                    if group_fields
                    else [data[metric.field].nunique(dropna=True)]
                )
            else:
                numeric = pd.to_numeric(data[metric.field], errors="coerce")
                operation = (
                    "mean" if metric.operation == "average" else metric.operation
                )
                if group_fields:
                    keyed = data[group_fields].copy()
                    keyed["__metric"] = numeric
                    metric_grouped = keyed.groupby(group_fields, dropna=False, sort=False)[
                        "__metric"
                    ]
                    values = getattr(metric_grouped, operation)().to_numpy()
                else:
                    values = [getattr(numeric, operation)()]
            output[metric.outputName] = values
            metric_columns.append(metric.outputName)
            if metric.ratioOutputName:
                numeric_output = pd.to_numeric(output[metric.outputName], errors="coerce")
                total = numeric_output.sum()
                output[metric.ratioOutputName] = (
                    numeric_output / total if total else 0
                )
                metric_columns.append(metric.ratioOutputName)
        output = output[[*group_fields, *metric_columns]]
        title = "分组统计结果"
        calculation = (
            f"筛选后按 {'、'.join(group_fields) or '全部记录'} 分组，"
            + "；".join(
                f"{metric.outputName}={metric.operation}"
                for metric in arguments.metrics
            )
            + "。"
        )
    else:
        fields = arguments.fields or [
            str(column) for column in data.columns[:30]
        ]
        output = data[fields]
        title = "查询结果"
        calculation = f"按 {len(arguments.filters)} 个条件筛选，并返回指定字段。"

    if arguments.sortBy and arguments.sortBy in output.columns:
        output = output.sort_values(
            arguments.sortBy,
            ascending=arguments.sortDirection == "asc",
            kind="stable",
        )
    output = output.head(arguments.limit)
    headers = [str(column) for column in output.columns]
    rows = [[_cell(value) for value in row] for row in output.itertuples(index=False, name=None)]
    return DataToolResult(
        requestId=request.id,
        tool="query_table",
        title=title,
        headers=headers,
        rows=rows,
        sourceSheets=[source.display_name for source in sources],
        scannedRows=scanned_rows,
        complete=not warnings,
        calculation=calculation,
        warnings=warnings,
    )
